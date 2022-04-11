
import pandas as pd
from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook

url="http://millenniumcricketleague.com/Home/Results.aspx?tt=1"

# Make a GET request to fetch the raw HTML content
html_content = requests.get(url).text

# Parse the html content
soup = BeautifulSoup(html_content, "lxml")
# print(soup.prettify()) # print the parsed data of html

x = soup.find_all('table', attrs={"class": "GridViewStyle"})[1]
print(x)

# data = {}
# t_headers = []
# for th in x.find_all("th"):
#     t_headers.append(th.text.replace('\n', ' ').strip())
# t_headers.append('samurl')
# print(t_headers)
# table_data = []

# for tr in x.find_all("tr"): # find all tr's from table's tbody
#         t_row = []
#         newrow=[]
#         for td in tr.find_all("td"):
#             newrow.append(td.text.replace('\n', ' ').strip())
#             for a in td.find_all('a', attrs={'class': 'Host'}):
#                 url = a['href']
#                 break
#         newrow.append(url)
#         table_data.append(newrow)

# print(table_data)
# df = pd.DataFrame(table_data)
# df.to_excel('ak2.xlsx')

#     # Put the data for the table with his heading.
# #
# # # # create a dummy list of texts to write to excel file
# # # divs = [[chr(i)*8, chr(i+1)*8] for i in range(65, 75, 1)]
# # #
# # # wb = Workbook()             # open new workbook, use load_workbook if existing
# # # ws = wb.create_sheet(title="Example")
# # # for div in divs:
# # #     row = [div[0], div[1]]  # construct a row: shown only for example purposes
# # #     ws.append(row)          # could use ws.append(div) since each div is a list
# # # wb.save('example.xlsx')