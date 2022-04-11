from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook
import pandas as pd
from openpyxl import load_workbook
url = "http://millenniumcricketleague.com/Home/Results.aspx?tt=1"
html_content = requests.get(url).text
soup = BeautifulSoup(html_content, "lxml")
x = soup.find_all('table', attrs={"class": "GridViewStyle"})[2]

stock = []

data = {}
# t_headers = []
# for th in x.find_all("th"):
#     t_headers.append(th.text.replace('\n', ' ').strip())
# t_headers.append('Links')
table_data = []
# table_data.append(t_headers)
#if False:


for tr in x.tbody.find_all("tr"):
    t_row = []
    newrow = []
    url = []
    urls = []
    i=0
    for td in tr.find_all("td"):
        newrow.append(td.text.replace('\n', ' ').strip())
        
        for a in td.find_all('a',{"href":True}):
            urls.append(a['href'])
    for i in urls[:1]:
        str1 = "http://millenniumcricketleague.com/Home/"
        i = str1 + i
        stock.append(i)
        newrow.append(i)
    table_data.append(newrow)
        # table_data_m = []
        # html_content = requests.get(i).text
        # soup = BeautifulSoup(html_content, "lxml")
        # x = soup.find_all('table')
        # for table in x[1:]:
        #     for tr2 in table.find_all("tr"):
        #         newrow2 = []
        #         for td2 in tr2.find_all("td"):
        #             newrow2.append(td2.text.replace('\n', ' ').strip())
        #         table_data_m.append(newrow2)
        #         print("Hi")

path = "F:\codes\show.xlsx"
writer = pd.ExcelWriter(path,engine='openpyxl')
counter = 1
# def custom_styles(val):
#     if val.name == 'Team Contacts':
#         return ['background-color: yellow'] 
                
for i in stock:
    print(i)
    table_data_m = []
    
    # for th in x.find_all("th"):
    #     t_headers.append(th.text.replace('\n', ' ').strip())
    #     table_data_m.append(t_headers)
    html_content = requests.get(i).text
    soup = BeautifulSoup(html_content, "lxml")
    x = soup.find('div', attrs={"id":"dvTeamDetails"})
    headings = []
    
    for hea in x.find_all("h3"):
        headings.append(hea.text.replace('\n', ' ').strip())
    head_count = 0
    for table in x.find_all("table"):
        joke =[]
        joke.append(headings[head_count])
        table_data_m.append(joke)
        
        head_count = head_count+1
        # table_data_m.append(headings[head_count].replace('\n', ' ').strip())
        # head_count = head_count+1
        if(table.find("th")):
            t_headers = []
            for th in table.find_all("th"):
                t_headers.append(th.text.replace('\n', ' ').strip())
            table_data_m.append(t_headers)
            
        for tr in table.find_all("tr"):
            newrow = []
            for td in tr.find_all("td"):
                newrow.append(td.text.replace('\n', ' ').strip())
                
                
            table_data_m.append(newrow)
            empty = []
            table_data_m.append(empty)
            



    table_data_m2 = [k for k in table_data_m if k!=[]]

    df = pd.DataFrame(table_data_m2)
    # df = df.style.apply(custom_styles)

    sheetname = 'Sheet%s' % counter
    if(counter!=1):
        writer.book = load_workbook(path)
    df.to_excel(writer, sheet_name=headings[0])
    print("Once!")
    counter = counter + 1
    writer.save()


    # df = pd.DataFrame(table_data_m)
    # sheetname = 'Sheet%s' % counter
    # if(counter!=1):
    #     writer.book = load_workbook(path)
    # df.to_excel(writer, sheet_name=sheetname)
    
    # counter = counter + 1
    # writer.save()
        
    # table_data.append(newrow)
# print(table_data)

# df = pd.DataFrame(table_data)
# df.to_excel('total_stats.xlsx')